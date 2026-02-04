import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter

st.title("Weekly Habanero Report Generator")

# --------------------------------------------------
# User inputs
# --------------------------------------------------
client = st.text_input("Client name (e.g. Wray's)")
report_number = st.text_input("Report number (e.g. 8)")

region = st.selectbox(
    "Client region",
    options=["US", "AU"],
    index=0
)

weekly_file = st.file_uploader(
    "Upload Weekly Data Pull",
    type=["xlsx"]
)

frequency_file = st.file_uploader(
    "Upload Frequency File",
    type=["xlsx"]
)

# --------------------------------------------------
# Run processing
# --------------------------------------------------
if st.button("Generate Report"):

    if not all([client, report_number, weekly_file, frequency_file]):
        st.error("Please complete all fields and upload both files.")
        st.stop()

    # ----------------------------------------------
    # Read Weekly
    # ----------------------------------------------
    weekly = pd.read_excel(weekly_file, sheet_name="Data")
    weekly["Date"] = pd.to_datetime(
        weekly["Date"], errors="coerce", format="mixed"
    )
    weekly = weekly.dropna(subset=["Date"])

    raw_start = weekly["Date"].min()
    raw_end = weekly["Date"].max()

    # ----------------------------------------------
    # Region-aware date logic
    # ----------------------------------------------
    if region == "US":
        start_date = raw_start - pd.Timedelta(days=1)
        end_date = raw_end - pd.Timedelta(days=1)
    else:  # AU
        start_date = raw_start
        end_date = raw_end

    if start_date.month == end_date.month:
        date_range = f"{start_date.strftime('%b %-d')} - {end_date.strftime('%-d')}"
    else:
        date_range = (
            f"{start_date.strftime('%b %-d')} - "
            f"{end_date.strftime('%b %-d')}"
        )

    # ----------------------------------------------
    # Read + clean Frequency
    # ----------------------------------------------
    freq = pd.read_excel(frequency_file, sheet_name="Data")

    freq["Date"] = pd.to_datetime(
        freq["Date"], errors="coerce", format="mixed"
    )
    freq = freq.dropna(subset=["Date"])

    freq = freq.rename(
        columns={"Unique Reach: Average Impression Frequency": "Frequency"}
    )

    freq["Frequency"] = pd.to_numeric(
        freq["Frequency"], errors="coerce"
    )

    # ----------------------------------------------
    # Merge + math
    # ----------------------------------------------
    merged = weekly.merge(
        freq[["Date", "Frequency"]],
        on="Date",
        how="left"
    )

    merged["Impressions"] = pd.to_numeric(
        merged["Impressions"], errors="coerce"
    )

    merged["Frequency"] = pd.to_numeric(
        merged["Frequency"], errors="coerce"
    )

    merged["Reach"] = merged["Impressions"] / merged["Frequency"]

    # ----------------------------------------------
    # Display table
    # ----------------------------------------------
    st.subheader(f"Preview — {client} ({date_range})")
    st.dataframe(merged, use_container_width=True)

    # ----------------------------------------------
    # Prepare Excel download (FORMAT ONLY)
    # ----------------------------------------------
    output_name = f"({report_number}) {client} {date_range}.xlsx"

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.book["Data"]

        # Column indices (1-based)
        date_col = merged.columns.get_loc("Date") + 1
        ctr_col = merged.columns.get_loc("Click Rate (CTR)") + 1
        reach_col = merged.columns.get_loc("Reach") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Date format
            row[date_col - 1].number_format = "yyyy/mm/dd"

            # CTR percentage
            row[ctr_col - 1].number_format = "0.00%"

            # Reach whole-number display
            row[reach_col - 1].number_format = "#,##0"

        # ------------------------------------------
        # Auto-adjust column widths
        # ------------------------------------------
        for col_idx, col_name in enumerate(merged.columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(col_name))

            for cell in ws[column_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

    st.download_button(
        label="Download Excel File",
        data=buffer.getvalue(),
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
